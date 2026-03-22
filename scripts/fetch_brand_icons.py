"""
fetch_brand_icons.py
--------------------
Reads the voiceover_jobs_list.xlsx, extracts all unique client/brand names,
downloads their logos from the Clearbit Logo API, saves them to
Images/Brand_Icons/, and produces a combined horizontal logo strip
(Images/Brand_Icons/_logo_strip.png) suitable for a "Trusted by" section.

Usage:
    py scripts/fetch_brand_icons.py

Re-run any time the xlsx is updated to refresh logos and the strip.
Requires: requests, Pillow, openpyxl
    py -m pip install requests Pillow openpyxl
"""

import os
import sys
import io
import requests
import openpyxl
from PIL import Image, ImageOps

# Fix Windows console encoding so non-ASCII brand names print correctly
if sys.stdout.encoding and sys.stdout.encoding.lower() not in ('utf-8', 'utf8'):
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

XLSX_PATH = os.path.join(os.path.dirname(__file__), '..', 'voiceover_jobs_list.xlsx')
OUTPUT_DIR = os.path.join(os.path.dirname(__file__), '..', 'Images', 'Brand_Icons')

# Height in pixels that all logos will be normalised to in the strip
STRIP_LOGO_HEIGHT = 80
# Horizontal padding between logos in the strip (px each side)
STRIP_LOGO_PADDING = 40
# Vertical padding above/below logos in the strip
STRIP_VERTICAL_PADDING = 30
# Background colour of the combined strip (RGBA)
STRIP_BG_COLOUR = (250, 248, 245, 255)  # matches site ivory #FAF8F5

# Map brand names (as they appear in the xlsx 'Client / Brand' column) to their
# website domain so Clearbit can find the right logo.
# Add new entries here when new clients are added to the xlsx.
DOMAIN_MAP = {
    'Migros':                         'migros.ch',
    'UBS':                            'ubs.com',
    'Coop':                           'coop.ch',
    'Swisscom':                       'swisscom.ch',
    'Swiss International Air Lines':  'swiss.com',
    'Nestlé':                         'nestle.com',
    'Lindt':                          'lindt.com',
    'Nespresso':                      'nespresso.com',
    'SBB CFF FFS':                    'sbb.ch',
    'PostFinance':                    'postfinance.ch',
    'Rolex':                          'rolex.com',
    'EPFL':                           'epfl.ch',
    'Raiffeisen':                     'raiffeisen.ch',
    'Omega':                          'omegawatches.com',
    'Helvetia Assurances':            'helvetia.com',
    'BCV (Banque Cantonale Vaudoise)':'bcv.ch',
    'Manor':                          'manor.ch',
    'Sunrise':                        'sunrise.ch',
    'La Mobilière':                   'mobiliere.ch',
    'CHUV':                           'chuv.ch',
}

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def get_brands_from_xlsx(path):
    """Return a list of unique brand names from the xlsx, preserving order."""
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    try:
        col_idx = headers.index('Client / Brand')
    except ValueError:
        sys.exit("ERROR: Could not find 'Client / Brand' column in xlsx.")
    seen = set()
    brands = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        brand = row[col_idx]
        if brand and brand not in seen:
            seen.add(brand)
            brands.append(brand)
    return brands


def fetch_logo(brand, domain, output_dir):
    """
    Download logo for the given domain, trying multiple public sources in order:
      1. icon.horse  — high-quality brand icons, no API key needed
      2. Favicon fallback via Google's favicon service (lower quality, last resort)
    Saves as <brand_slug>.png in output_dir.
    Returns the file path, or None on failure.
    """
    slug = brand.lower().replace(' ', '_').replace('(', '').replace(')', '').replace('/', '_')
    out_path = os.path.join(output_dir, f'{slug}.png')

    if os.path.exists(out_path):
        print(f'  [cached] {brand}')
        return out_path

    sources = [
        f'https://icon.horse/icon/{domain}',
        f'https://www.google.com/s2/favicons?domain={domain}&sz=256',
    ]

    headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'}

    for url in sources:
        try:
            resp = requests.get(url, timeout=15, headers=headers)
            if resp.status_code == 200 and 'image' in resp.headers.get('Content-Type', ''):
                img = Image.open(io.BytesIO(resp.content)).convert('RGBA')
                # Skip tiny favicons that are clearly low-quality (< 64x64)
                if img.width < 32 or img.height < 32:
                    continue
                img.save(out_path, 'PNG')
                print(f'  [ok]     {brand} -> {slug}.png  ({img.width}x{img.height})')
                return out_path
        except Exception as e:
            print(f'  [warn]   {brand} source {url} failed: {e}')
            continue

    print(f'  [fail]   {brand} — all sources failed')
    return None


def make_logo_strip(logo_paths, output_path, logo_height, padding, v_padding, bg_colour):
    """
    Resize all logos to logo_height (preserving aspect ratio),
    then compose them side-by-side with padding into a single PNG strip.
    Logos with transparent backgrounds are composited onto the bg colour.
    """
    images = []
    for path in logo_paths:
        try:
            img = Image.open(path).convert('RGBA')
            # Remove any pure-white background (common in downloaded logos)
            # by treating it as transparent — skip for logos that are already transparent
            aspect = img.width / img.height
            new_w = int(logo_height * aspect)
            img = img.resize((new_w, logo_height), Image.LANCZOS)
            images.append(img)
        except Exception as e:
            print(f'  [strip skip] {path} — {e}')

    if not images:
        print('No images to combine.')
        return

    total_width = sum(img.width + padding * 2 for img in images)
    total_height = logo_height + v_padding * 2

    strip = Image.new('RGBA', (total_width, total_height), bg_colour)

    x = 0
    for img in images:
        x += padding
        y = v_padding
        strip.paste(img, (x, y), img)
        x += img.width + padding

    # Flatten to RGB for PNG with white-ish bg
    final = Image.new('RGB', strip.size, bg_colour[:3])
    final.paste(strip, mask=strip.split()[3])
    final.save(output_path, 'PNG', optimize=True)
    print(f'\nStrip saved -> {output_path}')
    print(f'Dimensions: {final.width} x {final.height} px, {len(images)} logos')


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    print(f'Reading brands from: {os.path.abspath(XLSX_PATH)}')
    brands = get_brands_from_xlsx(XLSX_PATH)
    print(f'Found {len(brands)} unique brands: {", ".join(brands)}\n')

    logo_paths = []
    missing_domains = []

    for brand in brands:
        domain = DOMAIN_MAP.get(brand)
        if not domain:
            print(f'  [skip]   {brand} — no domain mapping (add to DOMAIN_MAP in this script)')
            missing_domains.append(brand)
            continue
        path = fetch_logo(brand, domain, OUTPUT_DIR)
        if path:
            logo_paths.append(path)

    print(f'\nDownloaded {len(logo_paths)} logos.')

    if missing_domains:
        print(f'\nBrands with no domain mapping (not included in strip):\n  ' + '\n  '.join(missing_domains))
        print('Add them to the DOMAIN_MAP dict at the top of this script and re-run.')

    if logo_paths:
        strip_path = os.path.join(OUTPUT_DIR, '_logo_strip.png')
        print('\nBuilding combined logo strip...')
        make_logo_strip(
            logo_paths,
            strip_path,
            logo_height=STRIP_LOGO_HEIGHT,
            padding=STRIP_LOGO_PADDING,
            v_padding=STRIP_VERTICAL_PADDING,
            bg_colour=STRIP_BG_COLOUR,
        )


if __name__ == '__main__':
    main()
