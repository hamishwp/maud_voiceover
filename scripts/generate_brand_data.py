#!/usr/bin/env python3
"""
generate_brand_data.py
Reads voiceover_jobs_list.xlsx and outputs js/brand-data.js.

js/brand-data.js is loaded by clients.html to populate the brand popup boxes.

Run this script whenever you update the spreadsheet:
    py scripts/generate_brand_data.py

The website will automatically reflect any changes on the next page load.
"""

import sys
import os
import json

# Windows encoding fix
if sys.stdout.encoding and sys.stdout.encoding.lower() not in ('utf-8', 'utf8'):
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

try:
    import openpyxl
except ImportError:
    print("openpyxl not found. Install it with: py -m pip install openpyxl")
    sys.exit(1)

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
ROOT_DIR   = os.path.dirname(SCRIPT_DIR)
XLSX_PATH  = os.path.join(ROOT_DIR, 'voiceover_jobs_list.xlsx')
OUTPUT_PATH = os.path.join(ROOT_DIR, 'js', 'brand-data.js')


def find_col(headers, *keywords):
    """Return the index of the first header that contains any of the keywords (case-insensitive)."""
    for i, h in enumerate(headers):
        if h is None:
            continue
        h_lower = str(h).lower()
        if any(kw.lower() in h_lower for kw in keywords):
            return i
    return None


def cell_str(row, idx):
    """Return a stripped string value from a row cell, or None."""
    if idx is None:
        return None
    val = row[idx]
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def main():
    if not os.path.exists(XLSX_PATH):
        print(f"ERROR: Spreadsheet not found at {XLSX_PATH}")
        sys.exit(1)

    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb.active

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    print(f"Columns: {headers}")

    # Locate columns by keyword — tolerant of header wording changes
    idx_brand    = find_col(headers, 'client', 'brand')
    idx_project  = find_col(headers, 'project')
    idx_language = find_col(headers, 'language', 'lang')
    idx_category = find_col(headers, 'category', 'type')
    idx_desc     = find_col(headers, 'description', 'desc')
    idx_duration = find_col(headers, 'duration', 'sec')

    print(f"Mapped columns -> brand:{idx_brand} project:{idx_project} "
          f"language:{idx_language} category:{idx_category} "
          f"desc:{idx_desc} duration:{idx_duration}")

    brand_data = {}  # { "Migros": [ {job}, ... ], ... }

    for row in ws.iter_rows(min_row=2, values_only=True):
        brand = cell_str(row, idx_brand)
        if not brand:
            continue

        job = {}
        project  = cell_str(row, idx_project)
        language = cell_str(row, idx_language)
        category = cell_str(row, idx_category)
        desc     = cell_str(row, idx_desc)
        duration = cell_str(row, idx_duration)

        if project:  job['project']  = project
        if language: job['language'] = language
        if category: job['category'] = category
        if desc:     job['description'] = desc
        if duration: job['duration'] = duration

        if brand not in brand_data:
            brand_data[brand] = []
        brand_data[brand].append(job)

    # Write output JS file
    json_str = json.dumps(brand_data, ensure_ascii=False, indent=2)
    js_content = f"""/**
 * brand-data.js — AUTO-GENERATED. DO NOT EDIT MANUALLY.
 *
 * Source: voiceover_jobs_list.xlsx
 * Regenerate: py scripts/generate_brand_data.py
 *
 * This file is loaded by clients.html to populate the brand popup boxes.
 * Add or update jobs in the spreadsheet, re-run the script, and the website
 * will automatically show the new information.
 */
window.BRAND_DATA = {json_str};
"""

    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
        f.write(js_content)

    print(f"\nWritten {len(brand_data)} brands to {OUTPUT_PATH}")
    for brand, jobs in brand_data.items():
        print(f"  {brand}: {len(jobs)} job(s)")


if __name__ == '__main__':
    main()
